# -*- coding: utf-8 -*-

from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
import datetime as dt
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, TimeoutException, UnexpectedAlertPresentException
import re

def crawling_bunjang():
    url = 'https://m.bunjang.co.kr/'
    # openApi페이지 클릭
    write_wb = Workbook();
    write_ws = write_wb.active

    # 엑셀 파일 첫째 줄 입력 -> 칼럼명 정의하기
    write_ws.cell(1, 1, '데이터 출처')
    write_ws.cell(1, 2, '회사명')
    write_ws.cell(1, 3, '업로드 일시')
    write_ws.cell(1, 4, '매물 지역')
    write_ws.cell(1, 5, '가격')
    write_ws.cell(1, 6, '제품 상태')
    write_ws.cell(1, 7, '기종')
    write_ws.cell(1, 8, '메모리')

    # header랑 option이  있어야 크롤러를 봇으로 인식을 안함
    header = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5)\
                    AppleWebKit 537.36 (KHTML, like Gecko) Chrome",
              "Accept": "text/html,application/xhtml+xml,application/xml;\
                    q=0.9,imgwebp,*/*;q=0.8"}
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko")
    driver = webdriver.Chrome(executable_path='chromedriver', options=options)

    driver.maximize_window()  # 모바일과 웹 뷰에서 위치가 달라질 수 있어서 maximize 하고 진행
    driver.get(url)  # 브라우저 실행 or 재실행 (재실행 시 위치 초기화 필요) 67 line에서 진행
    driver.implicitly_wait(10)
    #    driver.find_element(By.XPATH, '//*[@id="serviceGroups"]/li[2]/button').click()  # 카테고리 클릭

    #    select = Select(driver.find_element(By.XPATH, '//*[@id="sortColByCheck"]'))
    #    select.select_by_visible_text('제목순')

    #driver.find_element(By.XPATH, '//*[@class="flea-market-article-link"]').click()

    search=['아이폰']
    removeList=['케이스', 'case', '필름', '보호필름', '강화유리', '방탄', '충전기', '케이블', '충전케이블', '강화', '유리', '그립톡',
                '교환', '바꾸', '바꿀', '바꿔', '구해', '삽니다', '사요', '원해요', '원함', '구함', '사봅니다', '구합', '삼', '교신', '구매']
    newProductList=['미개봉', '미사용', '새상품', '사용안함']
    memoryList=['64', '128', '256', '512']

    deviceList=[['Xs', 'xs', 'XS', 'xS'], ['Xr', 'xr', 'XR', 'xR'], ['11'], ['12'], ['13'], ['Se', 'se', 'SE', 'sE'], ['Pro', 'pro', '프로', 'PRO'], ['Max', 'max', '맥스', 'MAX'], ['Mini', 'mini', '미니', 'MINI']]

    cnt= 0  # 매물 개수
    row= 2  # 엑셀 row
    page = 1    # 현재 페이지

    driver.find_element(By.XPATH, "//*[@class='sc-hMqMXs cLfdog']").click()
    # element = driver.find_element(By.CLASS_NAME, 'sc-hMqMXs cLfdog')
    element = driver.find_element(By.XPATH,"//*[@class='sc-hMqMXs cLfdog']")
    element.send_keys(search[0])

    # driver.find_element()

    driver.find_element(By.XPATH,"//*[@class='sc-kEYyzF gfWKdx']").click()

    # 스마트폰 필터 선택
    # driver.find_element_by_xpath("//*[@class='sc-ecaExY iRJhdC']/a[1]").click()

    # 최신순
    # driver.find_element_by_xpath("//*[@class='sc-iIHSe eSBGxA']/a[2]").click()
    count=0
    k=0
    while True:
        if count >= 1000:
            break
        # 선택할 매물 경로 설정
        xpathhead = '//*[@class="sc-eopZyb cXHRlj"]/div['
        xpathmiddle = str(k + 1)
        xpathtail = ']/a'
        xpath = xpathhead + xpathmiddle + xpathtail

        count += 1
        k += 1
        flag = True

        # 현재 페이지 끝나면 다음 페이지로 이동
        if count % 100 == 1 and count > 100:
            page += 1
            if page % 10 == 1:
                driver.find_element(By.XPATH, "//*[@class='sc-drlKqa zHqrz']/a[12]").click()
            else:
                driver.find_element(By.XPATH, '//a[text()=' + str(page) + ']').click()
            # wait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'more-btn'))).click()
            #
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@class='sc-eopZyb cXHRlj']")))
            print("다음페이지로 이동!")
            k = 0
            continue

        #지역
        try:
            region = driver.find_element(By.XPATH, xpath+'/div[3]').text
            region= region.split()[0]
            if '특별시' in region or '광역시' in region:
                region = region.rstrip('특별시')
                region = region.rstrip('광역시')
            elif '세종' in region:
                region = '세종시'
            elif '충청남도' in region:
                region = '충남'
            elif '충청북도' in region:
                region = '충북'
            elif '전라북도' in region:
                region = '전북'
            elif '전라남도' in region:
                region = '전남'
            elif '경상북도' in region:
                region = '경북'
            elif '경상남도' in region:
                region = '경남'
        except:
            region = '전국'


        # 매물상세페이지로 이동
        # count += 1
        # k += 1
        # flag = True
        try:
            driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
        except NoSuchElementException:
            print("NoSuchElementException 예외처리")
            continue


        try:
            # 가져온 데이터를 각 변수에 저장
            title = driver.find_element(By.XPATH,"//*[@class='sc-iFUGim igvOwa']").text  # 제목

            content = driver.find_element(By.XPATH,"//*[@class='sc-dOkuiw bphZvH']").text  # 내용

            # 날짜 계산
            current_time= dt.datetime.now().date()
            date = driver.find_element(By.XPATH, "//*[@class='sc-cNnxps QBKRV']/div[3]").text  # 날짜

            if '초' in date:
                t = int(date.rstrip('초 전'))
                date = current_time
            elif '분' in date:
                t= int(date.rstrip('분 전'))
                date= current_time - dt.timedelta(minutes=t)
            elif '시간' in date:
                t = int(date.rstrip('시간 전'))
                date = current_time - dt.timedelta(hours=t)
            elif '일' in date:
                t = int(date.rstrip('일 전'))
                date = current_time - dt.timedelta(days=t)
            elif '주' in date:
                t = int(date.rstrip('주 전'))
                t *= 7
                date = current_time - dt.timedelta(days=t)

            # 가격
            price = driver.find_element(By.XPATH, '//*[@class="sc-cNQqM hSmYTm"]/div[1]').text  # 가격
            if price == '가격없음' or price == '나눔' or re.search('\d', price)==None:
                price= 0
            else:
                price = int(price[:-1].replace(',', ''))

            # 건너뛰기
            # 제목에 케이스, 필름 들어간 거 건너뛰기
            for j in removeList:
                if j in title: # 케이스, 필름
                    flag= False

            # 가격 범위
            if price <= 70000 or price >= 2000000:
                flag= False

            # 메모리 정의
            memory = '없음'
            for j in memoryList:
                if j in title or j in content:
                    memory = j + 'GB'

            # 기종 정의
            if '128' in title:
                title= title.replace('128', '')
            elif '512' in title:
                title= title.replace('512', '')

            device = '아이폰'
            for r in deviceList:
                for j in r:
                    if j in title:
                        device += r[0]
                        break
            if device == '아이폰':
                flag = False
        except:
            print('예외 발생! 어디선가 뭐가 없네요ㅜ 다음 매물로 이동')
            driver.back()
            continue

        # 건너뛰어야하는 데이터는 처리
        if flag == False:
            driver.back()
            continue

        # 제품 상태 정의
        condition= False
        for m in newProductList:
            if m in title or m in content:
                condition=True


        print(str(count) + " " + title + "/" + str(date) + "/" + device + "/" + memory + "/" +region)

        write_ws.cell(row, 1, '번개장터')
        write_ws.cell(row, 2, '애플')
        write_ws.cell(row, 3, date)
        write_ws.cell(row, 4, region)
        write_ws.cell(row, 5, price)
        write_ws.cell(row, 6, condition)
        write_ws.cell(row, 7, device)
        write_ws.cell(row, 8, memory)

        # 뒤로가기
        driver.back()

        cnt += 1
        row += 1


    # driver.back()
    # driver.find_element(By.XPATH,"//*[@class='sc-hMqMXs cLfdog']").clear()   #검색창 비우기

    write_wb.save('번개장터_220317.xlsx')

    # 브라우저 종료
    driver.quit()

# crawling_bunjang()