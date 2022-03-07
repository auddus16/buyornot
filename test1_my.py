# -*- coding: utf-8 -*-

from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait

from selenium.common.exceptions import NoSuchElementException, TimeoutException, UnexpectedAlertPresentException

url = 'https://www.daangn.com/search/%EC%95%84%EC%9D%B4%ED%8F%B0'
# openApi페이지 클릭
write_wb = Workbook();
write_ws = write_wb.active

# 엑셀 파일 첫째 줄 입력 -> 칼럼명 정의하기
write_ws.cell(1, 1, '데이터 출처')
write_ws.cell(1, 2, '회사명')
write_ws.cell(1, 3, '업로드 일시')
write_ws.cell(1, 4, '매물 지역')
write_ws.cell(1, 5, '가격')
write_ws.cell(1, 6, '매물 제목')
write_ws.cell(1, 7, '매물 내용')

# header랑 option이  있어야 크롤러를 봇으로 인식을 안함
header = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5)\
    			AppleWebKit 537.36 (KHTML, like Gecko) Chrome",
              "Accept": "text/html,application/xhtml+xml,application/xml;\
    			q=0.9,imgwebp,*/*;q=0.8"}
options = webdriver.ChromeOptions()
# options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko")
driver = webdriver.Chrome(executable_path='chromedriver', options=options)

driver.maximize_window()  # 모바일과 웹 뷰에서 위치가 달라질 수 있어서 maximize 하고 진행
driver.get(url)  # 브라우저 실행 or 재실행 (재실행 시 위치 초기화 필요) 67 line에서 진행
driver.implicitly_wait(3)
#    driver.find_element(By.XPATH, '//*[@id="serviceGroups"]/li[2]/button').click()  # 카테고리 클릭

#    select = Select(driver.find_element(By.XPATH, '//*[@id="sortColByCheck"]'))
#    select.select_by_visible_text('제목순')

#driver.find_element(By.XPATH, '//*[@class="flea-market-article-link"]').click()

cnt= 0  # 매물 개수
row= 2  # 엑셀 row

while(cnt < 100): # 매물 100개 크롤링

    # 선택할 매물 경로 설정
    xpathhead = '//*[@id="flea-market-wrap"]/article['
    xpathmiddle = str(cnt + 1)
    xpathtail = ']/a'
    xpath = xpathhead + xpathmiddle + xpathtail

    # 매물상세페이지로 이동
    driver.find_element(By.XPATH, xpath).click()

   # 가져온 데이터를 각 변수에 저장
    region = driver.find_element(By.ID, 'region-name').text  # 지역
    title = driver.find_element(By.ID, 'article-title').text  # 제목
    content = driver.find_element(By.ID, 'article-detail').text  # 내용
    date = driver.find_element(By.TAG_NAME, 'time').text  # 날짜
    price = driver.find_element(By.XPATH, '//*[@id="article-description"]/p[5]').text  # 가격

    print(str(cnt)+title+region)

    write_ws.cell(row, 1, '당근마켓')
    write_ws.cell(row, 2, '애플')
    write_ws.cell(row, 3, date)
    write_ws.cell(row, 4, region)
    write_ws.cell(row, 5, price)
    write_ws.cell(row, 6, title)
    write_ws.cell(row, 7, content)

    # 뒤로가기
    driver.back()

    cnt+= 1
    row+= 1

    if cnt%6 == 0:
        wait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'more-btn'))).click()
        print("더보기 버튼 누름")

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'flea-market-wrap')))
        print("더보기 로드 완료")

write_wb.save('당근마켓_t1.xlsx')

# 브라우저 종료
driver.quit()

