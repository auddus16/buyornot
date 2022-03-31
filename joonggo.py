from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
import datetime as dt
import csv
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, TimeoutException, UnexpectedAlertPresentException
import time

url = 'https://m.joongna.com/search-list/product?cafeOrder=1&searchword=%EC%95%84%EC%9D%B4%ED%8F%B0&rid=HHaSE16CXCtuqLt'
# openApi페이지 클릭
write_wb = Workbook()
write_ws = write_wb.active

# 엑셀 파일 첫째 줄 입력 -> 칼럼명 정의하기
write_ws.cell(1, 1, '데이터 출처')
write_ws.cell(1, 2, '회사명')
write_ws.cell(1, 3, '업로드 일시')
write_ws.cell(1, 4, '매물 지역')
write_ws.cell(1, 5, '가격')
write_ws.cell(1, 6, '제품 상태')
write_ws.cell(1, 7, '기종')
write_ws.cell(1, 8, '메모리')

# header랑 option이  있어야 크롤러를 봇으로 인식을 안함
header = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5)\
    			AppleWebKit 537.36 (KHTML, like Gecko) Chrome",
          "Accept": "text/html,application/xhtml+xml,application/xml;\
    			q=0.9,imgwebp,*/*;q=0.8"}
options = webdriver.ChromeOptions()
# options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko")
driver = webdriver.Chrome(executable_path='chromedriver', options=options)

driver.maximize_window()  # 모바일과 웹 뷰에서 위치가 달라질 수 있어서 maximize 하고 진행
driver.get(url)  # 브라우저 실행 or 재실행 (재실행 시 위치 초기화 필요) 67 line에서 진행
driver.implicitly_wait(30)


search = ['아이폰']
removeList = ['케이스', 'case', '필름', '보호필름', '강화유리', '방탄', '충전기', '케이블', '충전케이블', '강화', '유리', '그립톡',
              '교환', '바꾸', '바꿀', '바꿔', '구해', '삽니다', '사요', '원해요', '원함', '구함', '사봅니다', '구합', '삼', '교신', '구매']
newProductList = ['미개봉', '미사용', '새상품', '사용안함']
memoryList = ['64', '128', '256', '512']

deviceList = [['Xs', 'xs', 'XS', 'xS'], ['11'], ['Xr', 'xr', 'XR', 'xR'], ['12'], ['13'], ['Se', 'se', 'SE', 'sE'],
              ['Pro', 'pro', '프로', 'PRO'], ['Max', 'max', '맥스', 'MAX'], ['Mini', 'mini', '미니', 'MINI']]

cnt = 0  # 매물 개수
row = 2  # 엑셀 row


for k in range(130):
    # time.sleep(1)
    
# 선택할 매물 경로 설정
    xpathhead = '//*[@class="SearchList_listWrap__14Cu9 pd_h20"]/div/div['
    xpathmiddle = str(k + 1)
    xpathtail = ']/div/a'
    xpath = xpathhead + xpathmiddle + xpathtail
    

    region = '지역없음'

    driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)


    flag = True

    try:
        # 가져온 데이터를 각 변수에 저장
        title = driver.find_element(By.XPATH, '//*[@class="f18 ProductDetailComponent_title__20tny"]').text  # 제목
        content = driver.find_element(By.XPATH, '//*[@class="pd_b30"]').text  # 내용
        date = driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div/p[1]').text  # 날짜
        price = driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/p/strong').text  # 가격


        # 날짜 계산
        current_time = dt.datetime.now().date()

        if '초' in date:
            t = int(date.rstrip('초 전'))
            date = current_time
        elif '분' in date:
            t= int(date.rstrip('분 전'))
            date= current_time - dt.timedelta(minutes=t)
        elif '시간' in date:
            t = int(date.rstrip('시간 전'))
            date = current_time - dt.timedelta(hours=t)
        elif '일' in date:
            t = int(date.rstrip('일 전'))
            date = current_time - dt.timedelta(days=t)
        elif '주' in date:
            t = int(date.rstrip('주 전'))
            t *= 7
            date = current_time - dt.timedelta(days=t)

        price = int(price[:-1].replace(',', ''))

        # 제목에 케이스, 필름 들어간 거 건너뛰기
        for j in removeList:
            if j in title:  # 케이스, 필름
                flag = False

        # 가격 범위
        if price <= 70000 or price >= 2000000:
            flag = False

        # 메모리 정의
        memory = '없음'
        for j in memoryList:
            if j in title or j in content:
                memory = j + 'GB'

        # 기종 정의
        if '128' in title:
            title = title.replace('128', '')
        elif '512' in title:
            title = title.replace('512', '')

        device = '아이폰'
        for r in deviceList:
            for j in r:
                if j in title:
                    device += r[0]
                    break
        if device == '아이폰':
            flag = False               
        
    except:
        print('예외 발생! 어디선가 뭐가 없네요ㅜ 다음 매물로 이동')
        driver.back()
        continue

    # 건너뛰어야하는 데이터는 처리
    if flag == False:
        driver.back()
        if (k + 1) % 10 == 0:
            wait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/section/article/button'))).click()

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div[1]/section/article')))
        continue

    # 제품 상태 정의
    condition = False
    for m in newProductList:
        if m in title or m in content:
            condition = True

    print(str(k) + " " + title + "/" + "/" + str(date) + "/" + region)

    write_ws.cell(row, 1, '중고나라')
    write_ws.cell(row, 2, '애플')
    write_ws.cell(row, 3, date)
    write_ws.cell(row, 4, region)
    write_ws.cell(row, 5, price)
    write_ws.cell(row, 6, condition)
    write_ws.cell(row, 7, device)
    write_ws.cell(row, 8, memory)

    # 뒤로가기
    driver.back()

    cnt += 1
    row += 1

    if (k + 1) % 10 == 0:
        wait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/section/article/button'))).click()

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div[1]/section/article')))

# 데이터 엑셀에 저장
write_wb.save('중고나라_220331_1.xlsx')

# 브라우저 종료
driver.quit()
