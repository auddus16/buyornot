from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
import datetime as dt
import csv
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, TimeoutException, UnexpectedAlertPresentException
from sklearn.tree import DecisionTreeClassifier

url = 'https://m.joongna.com/search-list/product?cafeOrder=1&searchword=%EC%95%84%EC%9D%B4%ED%8F%B0&rid=HHaSE16CXCtuqLt'
# openApi페이지 클릭
write_wb = Workbook();
write_ws = write_wb.active

# 엑셀 파일 첫째 줄 입력 -> 칼럼명 정의하기
write_ws.cell(1, 1, '데이터 출처')
write_ws.cell(1, 2, '회사명')
write_ws.cell(1, 3, '업로드 일시')
write_ws.cell(1, 4, '매물 지역')
write_ws.cell(1, 5, '가격')
write_ws.cell(1, 6, '제품 상태')
write_ws.cell(1, 7, '기종')
write_ws.cell(1, 8, '메모리')

# header랑 option이  있어야 크롤러를 봇으로 인식을 안함
header = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5)\
    			AppleWebKit 537.36 (KHTML, like Gecko) Chrome",
          "Accept": "text/html,application/xhtml+xml,application/xml;\
    			q=0.9,imgwebp,*/*;q=0.8"}
options = webdriver.ChromeOptions()
# options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko")
driver = webdriver.Chrome(executable_path='chromedriver', options=options)

driver.maximize_window()  # 모바일과 웹 뷰에서 위치가 달라질 수 있어서 maximize 하고 진행
driver.get(url)  # 브라우저 실행 or 재실행 (재실행 시 위치 초기화 필요) 67 line에서 진행
driver.implicitly_wait(10)


search = ['아이폰']
removeList = ['케이스', 'case', '필름', '보호필름', '강화유리', '방탄', '충전기', '케이블', '충전케이블', '강화', '유리', '그립톡',
              '교환', '바꾸', '바꿀', '바꿔', '구해', '삽니다', '사요', '원해요', '원함', '구함', '사봅니다', '구합', '삼', '교신', '구매']
newProductList = ['미개봉', '미사용', '새상품', '사용안함']
memoryList = ['64', '128', '256', '512']

deviceList = [['Xs', 'xs', 'XS', 'xS'], ['11'], ['Xr', 'xr', 'XR', 'xR'], ['12'], ['13'], ['Se', 'se', 'SE', 'sE'],
              ['Pro', 'pro', '프로', 'PRO'], ['Max', 'max', '맥스', 'MAX'], ['Mini', 'mini', '미니', 'MINI']]

cnt = 0  # 매물 개수
row = 2  # 엑셀 row


for k in range(200):
    
# 선택할 매물 경로 설정
    xpathhead = '//*[@class="SearchList_listWrap__14Cu9 pd_h20"]/div/div['
    xpathmiddle = str(k + 1)
    xpathtail = ']div/a'
    xpath = xpathhead + xpathmiddle + xpathtail
    

    region = '전국'

    try:
        driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
    except NoSuchElementException:  # 선택할 수 있는 매물이 없는 경우, 더보기 버튼 누르기->처음으로
        print('예외 발생! 더보기 버튼 누르기')
        wait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'SearchList_moreButton__11RNU'))).click()

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'SearchList_listWrap__14Cu9 pd_h20')))
        continue

    # 불필요한 데이터 건너뛰기 위한 변수 정의 -> 전처리
    flag = True

    try:
        # 가져온 데이터를 각 변수에 저장
        title = driver.find_element(By.CLASS_NAME, 'f18 ProductDetailComponent_title__20tny').text  # 제목

        content = driver.find_element(By.CLASS_NAME, 'pd_b30').text  # 내용

        date = driver.find_element(By.XPATH, '//*[@class="ProductDetailComponent_etc__3Vh-z f14 c_gray mt10"]/p[1]').text  # 날짜

        price = driver.find_element(By.CLASS_NAME, 'ProductDetailComponent_price__17OTs').text  # 가격

        
    except:
        print('예외 발생! 어디선가 뭐가 없네요ㅜ 다음 매물로 이동')
        driver.back()
        continue

    if flag == False:
        driver.back()
        if (k + 1) % 11 == 0:
            wait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'SearchList_moreButton__11RNU'))).click()

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'SearchList_listWrap__14Cu9 pd_h20')))
        continue

    # 제품 상태 정의
    condition = False
    for m in newProductList:
        if m in title or m in content:
            condition = True

    print(str(k) + " " + title + "/" + "/" + str(date) + "/" + region)

    write_ws.cell(row, 1, '중고나라')
    write_ws.cell(row, 2, '애플')
    write_ws.cell(row, 3, date)
    write_ws.cell(row, 4, region)
    write_ws.cell(row, 5, price)
    write_ws.cell(row, 6, condition)
    #write_ws.cell(row, 7, device)
    #write_ws.cell(row, 8, memory)

    # 뒤로가기
    driver.back()

    cnt += 1
    row += 1

    if (k + 1) % 11 == 0:
        wait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'SearchList_moreButton__11RNU'))).click()

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'SearchList_listWrap__14Cu9 pd_h20')))

# 데이터 엑셀에 저장
write_wb.save('중고나라_t1.xlsx')

# 브라우저 종료
driver.quit()
